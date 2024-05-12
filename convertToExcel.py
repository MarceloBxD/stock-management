from database import Database
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.formatting.rule import Rule
from openpyxl.styles.differential import DifferentialStyle

def convert():
    # Create workbook
    wb = openpyxl.Workbook()
    # Get workbook active sheet
    sheet = wb.active
    # Creating Headings
    sheet.cell(row=1, column=1).value = "CR do Produto"
    sheet.cell(row=1, column=2).value = "Nome do Produto"
    sheet.cell(row=1, column=3).value = "Quantidade em Estoque"

    # Styling Headings
    for i in range(1, 5):
        sheet.cell(row=1, column=i).font = Font(bold=True, name="arial", size=15)
        sheet.cell(row=1, column=i).fill = PatternFill(patternType="solid", fgColor="00e5ee")
        sheet.cell(row=1, column=i).alignment = Alignment(horizontal="center")

    # Setting column dimensions
    for col in "ABCDEFGH":
        sheet.column_dimensions[col].width = 25

    # Retrieving date from database
    db = Database("products.db")
    rows = db.fetch_all_rows()

    # Filling date from Database into the worksheet
    for sheet_row, row in enumerate(rows):
        for sheet_col, item in enumerate(row):
            sheet.cell(row=sheet_row + 2, column=sheet_col + 1).value = item
            sheet.cell(row=sheet_row + 2, column=sheet_col + 1).alignment = Alignment(horizontal="center")

    wb.save("product_list.xlsx")


    # loading excel file to read and write to
    wb = openpyxl.load_workbook(filename="product_list.xlsx")

    sheet = wb.active

    wb.save("product_list.xlsx")



